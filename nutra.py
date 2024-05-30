import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import os
from dotenv import load_dotenv

load_dotenv()

KEY_TABLE = os.getenv('KEY_TABLE')

def rooo():

    wb5 = openpyxl.Workbook()
    sheet6 = wb5.active

    scope = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name("wbapi-408006-f5818dfd0b37.json", scope)
    client = gspread.authorize(credentials)

    spreadsheet_id = '1bjheBkVlfDmbcJbmK3OsTtbZQ0Sjwc5ZMuyXqvWRi5w'

    sheet_name = 'R%'

    sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

    all_data = sheet.get_all_values()

    row111 = all_data[115]
    row212 = all_data[120]
    row311 = all_data[121]
    row412 = all_data[122]
    row511 = all_data[123]
    row611 = all_data[124]

    row_b = all_data[157]
    row_c = all_data[162]
    row_d = all_data[163]
    row_e = all_data[164]
    row_f = all_data[165]
    row_g = all_data[166]

    row_b1 = all_data[199]
    row_c1 = all_data[204]
    row_d1 = all_data[205]
    row_e1 = all_data[206]
    row_f1 = all_data[207]
    row_g1 = all_data[208]

    row_b11 = all_data[241]
    row_c11 = all_data[246]
    row_d11 = all_data[247]
    row_e11 = all_data[248]
    row_f11 = all_data[249]
    row_g11 = all_data[250]

    row_b111 = all_data[283]
    row_c111 = all_data[288]
    row_d111 = all_data[289]
    row_e111 = all_data[290]
    row_f111 = all_data[291]
    row_g111 = all_data[292]

    row_b1111 = all_data[325]
    row_c1111 = all_data[330]
    row_d1111 = all_data[331]
    row_e1111 = all_data[332]
    row_f1111 = all_data[333]
    row_g1111 = all_data[334]

    row_datapu1 = row_b
    row_data231 = row_c
    row_data341 = row_d
    row_data451 = row_e
    row_data561 = row_f
    row_data671 = row_g

    row_datapu11 = row111
    row_data2311 = row212
    row_data3411 = row311
    row_data4511 = row412
    row_data5611 = row511
    row_data6711 = row611

    row_datapu111 = row_b1
    row_data23111 = row_c1
    row_data34111 = row_d1
    row_data45111 = row_e1
    row_data56111 = row_f1
    row_data67111 = row_g1

    row_datapu1111 = row_b11
    row_data231111 = row_c11
    row_data341111 = row_d11
    row_data451111 = row_e11
    row_data561111 = row_f11
    row_data671111 = row_g11

    row_datapu11111 = row_b111
    row_data2311111 = row_c111
    row_data3411111 = row_d111
    row_data4511111 = row_e111
    row_data5611111 = row_f111
    row_data6711111 = row_g111

    row_datapu111111 = row_b1111
    row_data23111111 = row_c1111
    row_data34111111 = row_d1111
    row_data45111111 = row_e1111
    row_data56111111 = row_f1111
    row_data67111111 = row_g1111

    my_list211 = [item.replace('\xa0', '') for item in row_datapu1]
    my_list311 = [item.replace('\xa0', '') for item in row_data231]
    my_list411 = [item.replace('\xa0', '') for item in row_data341]
    my_list511 = [item.replace('\xa0', '') for item in row_data451]
    my_list611 = [item.replace('\xa0', '') for item in row_data561]
    my_list711 = [item.replace('\xa0', '') for item in row_data671]

    my_list2111 = [item.replace('\xa0', '') for item in row_datapu11]
    my_list3111 = [item.replace('\xa0', '') for item in row_data2311]
    my_list4111 = [item.replace('\xa0', '') for item in row_data3411]
    my_list5111 = [item.replace('\xa0', '') for item in row_data4511]
    my_list6111 = [item.replace('\xa0', '') for item in row_data5611]
    my_list7111 = [item.replace('\xa0', '') for item in row_data6711]

    my_list21111 = [item.replace('\xa0', '') for item in row_datapu111]
    my_list31111 = [item.replace('\xa0', '') for item in row_data23111]
    my_list41111 = [item.replace('\xa0', '') for item in row_data34111]
    my_list51111 = [item.replace('\xa0', '') for item in row_data45111]
    my_list61111 = [item.replace('\xa0', '') for item in row_data56111]
    my_list71111 = [item.replace('\xa0', '') for item in row_data67111]

    my_list211111 = [item.replace('\xa0', '') for item in row_datapu1111]
    my_list311111 = [item.replace('\xa0', '') for item in row_data231111]
    my_list411111 = [item.replace('\xa0', '') for item in row_data341111]
    my_list511111 = [item.replace('\xa0', '') for item in row_data451111]
    my_list611111 = [item.replace('\xa0', '') for item in row_data561111]
    my_list711111 = [item.replace('\xa0', '') for item in row_data671111]

    my_list2111111 = [item.replace('\xa0', '') for item in row_datapu11111]
    my_list3111111 = [item.replace('\xa0', '') for item in row_data2311111]
    my_list4111111 = [item.replace('\xa0', '') for item in row_data3411111]
    my_list5111111 = [item.replace('\xa0', '') for item in row_data4511111]
    my_list6111111 = [item.replace('\xa0', '') for item in row_data5611111]
    my_list7111111 = [item.replace('\xa0', '') for item in row_data6711111]

    my_list21111111 = [item.replace('\xa0', '') for item in row_datapu111111]
    my_list31111111 = [item.replace('\xa0', '') for item in row_data23111111]
    my_list41111111 = [item.replace('\xa0', '') for item in row_data34111111]
    my_list51111111 = [item.replace('\xa0', '') for item in row_data45111111]
    my_list61111111 = [item.replace('\xa0', '') for item in row_data56111111]
    my_list71111111 = [item.replace('\xa0', '') for item in row_data67111111]

    #Organic Form
    for i in range(len(my_list211[266:297])):
        sheet6.cell(row=202, column=i + 4, value=my_list211[i+266])
    for i in range(len(my_list311[266:297])):
        sheet6.cell(row=203, column=i + 4, value=my_list311[i+266])
    for i in range(len(my_list411[266:297])):
        sheet6.cell(row=204, column=i + 4, value=my_list411[i+266])
    for i in range(len(my_list511[266:297])):
        sheet6.cell(row=205, column=i + 4, value=my_list511[i+266])
    for i in range(len(my_list611[266:297])):
        sheet6.cell(row=206, column=i + 4, value=my_list611[i+266])
    for i in range(len(my_list711[266:297])):
        sheet6.cell(row=207, column=i + 4, value=my_list711[i+266])

    #Nutrisystem
    for i in range(len(my_list2111[266:297])):
        sheet6.cell(row=138, column=i + 4, value=my_list2111[i+266])
    for i in range(len(my_list3111[266:297])):
        sheet6.cell(row=139, column=i + 4, value=my_list3111[i+266])
    for i in range(len(my_list4111[266:297])):
        sheet6.cell(row=140, column=i + 4, value=my_list4111[i+266])
    for i in range(len(my_list5111[266:297])):
        sheet6.cell(row=141, column=i + 4, value=my_list5111[i+266])
    for i in range(len(my_list6111[266:297])):
        sheet6.cell(row=142, column=i + 4, value=my_list6111[i+266])
    for i in range(len(my_list7111[266:297])):
        sheet6.cell(row=143, column=i + 4, value=my_list7111[i+266])

    #Orsogood
    for i in range(len(my_list21111[266:297])):
        sheet6.cell(row=506, column=i + 4, value=my_list21111[i+266])
    for i in range(len(my_list31111[266:297])):
        sheet6.cell(row=507, column=i + 4, value=my_list31111[i+266])
    for i in range(len(my_list41111[266:297])):
        sheet6.cell(row=508, column=i + 4, value=my_list41111[i+266])
    for i in range(len(my_list51111[266:297])):
        sheet6.cell(row=509, column=i + 4, value=my_list51111[i+266])
    for i in range(len(my_list61111[266:297])):
        sheet6.cell(row=510, column=i + 4, value=my_list61111[i+266])
    for i in range(len(my_list71111[266:297])):
        sheet6.cell(row=511, column=i + 4, value=my_list71111[i+266])

    #Orsogood 90 капсул
    for i in range(len(my_list211111[266:297])):
        sheet6.cell(row=778, column=i + 4, value=my_list211111[i+266])
    for i in range(len(my_list311111[266:297])):
        sheet6.cell(row=779, column=i + 4, value=my_list311111[i+266])
    for i in range(len(my_list411111[266:297])):
        sheet6.cell(row=780, column=i + 4, value=my_list411111[i+266])
    for i in range(len(my_list511111[266:297])):
        sheet6.cell(row=781, column=i + 4, value=my_list511111[i+266])
    for i in range(len(my_list611111[266:297])):
        sheet6.cell(row=782, column=i + 4, value=my_list611111[i+266])
    for i in range(len(my_list711111[266:297])):
        sheet6.cell(row=783, column=i + 4, value=my_list711111[i+266])

    #Perfect waist 90 капсул
    for i in range(len(my_list2111111[266:297])):
        sheet6.cell(row=970, column=i + 4, value=my_list2111111[i+266])
    for i in range(len(my_list3111111[266:297])):
        sheet6.cell(row=971, column=i + 4, value=my_list3111111[i+266])
    for i in range(len(my_list4111111[266:297])):
        sheet6.cell(row=972, column=i + 4, value=my_list4111111[i+266])
    for i in range(len(my_list5111111[266:297])):
        sheet6.cell(row=973, column=i + 4, value=my_list5111111[i+266])
    for i in range(len(my_list6111111[266:297])):
        sheet6.cell(row=974, column=i + 4, value=my_list6111111[i+266])
    for i in range(len(my_list7111111[266:297])):
        sheet6.cell(row=975, column=i + 4, value=my_list7111111[i+266])

    #Perfect waist
    for i in range(len(my_list21111111[266:297])):
        sheet6.cell(row=874, column=i + 4, value=my_list21111111[i+266])
    for i in range(len(my_list31111111[266:297])):
        sheet6.cell(row=875, column=i + 4, value=my_list31111111[i+266])
    for i in range(len(my_list41111111[266:297])):
        sheet6.cell(row=876, column=i + 4, value=my_list41111111[i+266])
    for i in range(len(my_list51111111[266:297])):
        sheet6.cell(row=877, column=i + 4, value=my_list51111111[i+266])
    for i in range(len(my_list61111111[266:297])):
        sheet6.cell(row=878, column=i + 4, value=my_list61111111[i+266])
    for i in range(len(my_list71111111[266:297])):
        sheet6.cell(row=879, column=i + 4, value=my_list71111111[i+266])

    wb5.save("sheet6.xlsx")

    def get_column_label(i):
        if i < 26:
            return chr(65 + i)
        else:
            div = i // 26
            mod = i % 26
            if mod == 0:
                return get_column_label(div - 1) + 'Z'
            else:
                return get_column_label(div - 1) + get_column_label(mod)

    def CopyFromExcInGsh5():
        client = gspread.authorize(credentials)

        spreadsheet = client.open_by_key(KEY_TABLE)
        worksheet = spreadsheet.worksheet('Аналитика и статистика все компании')

        df = pd.read_excel("sheet6.xlsx")
        data_list = df.values.tolist()
        num_cols = len(data_list[0])

        cell_list = worksheet.range('A1:' + get_column_label(num_cols - 1) + str(len(data_list)))
        for cell in cell_list:
            row = (cell.row - 1) if (cell.row - 1) < len(data_list) else -1
            col = (cell.col - 1) if (cell.col - 1) < num_cols else -1
            if row != -1 and col != -1:
                value = data_list[row][col]
                if pd.notna(value):
                    cell.value = str(value)

        worksheet.update_cells(cell_list)
        print("Данные загружены")

    CopyFromExcInGsh5()