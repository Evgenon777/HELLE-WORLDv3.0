import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

saved_positions = {}
columnStat = 3

def Data(source, dictionary, dates):

    start_row = 2
    start_column = 3
    column = 2
    row_ro = 3
    row_pi = 2

    for ID in source:
        sheet.cell(row=row_pi, column=column, value="НАЗВАНИЕ БРЕНДА")
        sheet.cell(row=row_pi, column=column - 1, value="МЕТРИКА")
        sheet.cell(row=row_pi, column=column + 1, value="Артикул")
        sheet.cell(row=row_ro, column=columnStat + 1, value="--------------------")
        sheet.cell(row=row_ro , column=columnStat - 2, value="--------------------")
        sheet.cell(row=row_ro , column=columnStat - 1, value="--------------------")
        sheet.cell(row=row_ro , column=columnStat, value="--------------------")

        row_pi += 15
        if ID in dictionary:
            data = dictionary[ID]

            if ID in saved_positions:
                row_ro = saved_positions[ID]
            else:
                row_ro += 1
                saved_positions[ID] = row_ro
            sheet.cell(row=row_ro, column=column - 1, value="Показы")
            sheet.cell(row=row_ro, column=columnStat + 1, value=data['Показы'])
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="Переходы")
            sheet.cell(row=row_ro, column=columnStat + 1, value=data['Переходы'])
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="CTR")
            sheet.cell(row=row_ro, column=columnStat + 1, value=data['CTR'])
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="Конверсии в корзину")
            sheet.cell(row=row_ro, column=columnStat + 1, value=str(data['Конверсии в корзину']) + "%")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="Конверсии в заказ")
            sheet.cell(row=row_ro, column=columnStat + 1, value=str(data['Конверсии в заказ']) + "%")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="Остатки товаров на складе")
            sheet.cell(row=row_ro, column=columnStat + 1, value=data['Остатки товаров на складе'])
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1

            sheet.cell(row=row_ro, column=column - 1, value="Валовая прибыль")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1

            sheet.cell(row=row_ro, column=column - 1, value="Чистая прибыль")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro+=1

            sheet.cell(row=row_ro, column=column - 1, value="R%")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro+=1
            sheet.cell(row=row_ro, column=column - 1, value="Ср.чек")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro+=1
            sheet.cell(row=row_ro, column=column - 1, value="РС")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro, column=column - 1, value="ДРР")
            sheet.cell(row=row_ro, column=column, value=data['Бренд'])
            sheet.cell(row=row_ro, column=column + 1, value=data['ID'])
            row_ro += 1
            sheet.cell(row=row_ro+1, column=columnStat + 1, value="--------------------")
            sheet.cell(row=row_ro+1, column=columnStat - 2, value="--------------------")
            sheet.cell(row=row_ro+1, column=columnStat - 1, value="--------------------")
            sheet.cell(row=row_ro+1, column=columnStat , value="--------------------")

            row_ro += 3
            row_pi += 1

        for i, date in enumerate(dates):
            sheet.cell(row=start_row, column=start_column + i + 1, value=date.strftime('%d.%m.%y'))
        start_row += 16

    wb.save("sheet.xlsx")