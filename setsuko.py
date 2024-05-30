import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import os
from dotenv import load_dotenv

load_dotenv()

KEY_TABLE = os.getenv('KEY_TABLE')

def rooo():

    wb5 = openpyxl.Workbook()
    sheet6 = wb5.active

    scope = ['https://www.googleapis.com/auth/spreadsheets']
    credentials = ServiceAccountCredentials.from_json_keyfile_name("wbapi-408006-f5818dfd0b37.json", scope)
    client = gspread.authorize(credentials)

    spreadsheet_id = '1bavcWS8fvhe5h7BeCDzyfJHp3RZhii3SIk0iOlWLmyM'

    sheet_name = 'R%'

    sheet = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

    all_data = sheet.get_all_values()

    row_index_74 = all_data[73]
    row_index = all_data[78]
    row_index_80 = all_data[79]
    row_index_81 = all_data[80]
    row_index_82 = all_data[81]
    row_index_83 = all_data[82]

    row1 = all_data[31]
    row2 = all_data[36]
    row3 = all_data[37]
    row4 = all_data[38]
    row5 = all_data[39]
    row6 = all_data[40]

    row111 = all_data[115]
    row212 = all_data[120]
    row311 = all_data[121]
    row412 = all_data[122]
    row511 = all_data[123]
    row611 = all_data[124]

    row_b = all_data[157]
    row_c = all_data[162]
    row_d = all_data[163]
    row_e = all_data[164]
    row_f = all_data[165]
    row_g = all_data[166]

    row_data = row_index
    row_data2 = row_index_74
    row_data3 = row_index_80
    row_data4 = row_index_81
    row_data5 = row_index_82
    row_data6 = row_index_83

    row_datapu = row1
    row_data23 = row2
    row_data34 = row3
    row_data45 = row4
    row_data56 = row5
    row_data67 = row6

    row_datapu1 = row_b
    row_data231 = row_c
    row_data341 = row_d
    row_data451 = row_e
    row_data561 = row_f
    row_data671 = row_g

    row_datapu11 = row111
    row_data2311 = row212
    row_data3411 = row311
    row_data4511 = row412
    row_data5611 = row511
    row_data6711 = row611

    my_list21 = [item.replace('\xa0', '') for item in row_datapu]
    my_list31 = [item.replace('\xa0', '') for item in row_data23]
    my_list41 = [item.replace('\xa0', '') for item in row_data34]
    my_list51 = [item.replace('\xa0', '') for item in row_data45]
    my_list61 = [item.replace('\xa0', '') for item in row_data56]
    my_list71 = [item.replace('\xa0', '') for item in row_data67]

    my_list2 = [item.replace('\xa0', '') for item in row_data]
    my_list3 = [item.replace('\xa0', '') for item in row_data2]
    my_list4 = [item.replace('\xa0', '') for item in row_data3]
    my_list5 = [item.replace('\xa0', '') for item in row_data4]
    my_list6 = [item.replace('\xa0', '') for item in row_data5]
    my_list7 = [item.replace('\xa0', '') for item in row_data6]

    my_list211 = [item.replace('\xa0', '') for item in row_datapu1]
    my_list311 = [item.replace('\xa0', '') for item in row_data231]
    my_list411 = [item.replace('\xa0', '') for item in row_data341]
    my_list511 = [item.replace('\xa0', '') for item in row_data451]
    my_list611 = [item.replace('\xa0', '') for item in row_data561]
    my_list711 = [item.replace('\xa0', '') for item in row_data671]

    my_list2111 = [item.replace('\xa0', '') for item in row_datapu11]
    my_list3111 = [item.replace('\xa0', '') for item in row_data2311]
    my_list4111 = [item.replace('\xa0', '') for item in row_data3411]
    my_list5111 = [item.replace('\xa0', '') for item in row_data4511]
    my_list6111 = [item.replace('\xa0', '') for item in row_data5611]
    my_list7111 = [item.replace('\xa0', '') for item in row_data6711]

    #enzyme powder
    for i in range(len(my_list3[167:198])):
        sheet6.cell(row=1242, column=i + 3, value=my_list3[i+167])
    for i in range(len(my_list2[167:198])):
        sheet6.cell(row=1243, column=i + 3, value=my_list2[i+167])
    for i in range(len(my_list4[167:198])):
        sheet6.cell(row=1244, column=i + 3, value=my_list4[i+167])
    for i in range(len(my_list5[167:198])):
        sheet6.cell(row=1245, column=i + 3, value=my_list5[i+167])
    for i in range(len(my_list6[167:198])):
        sheet6.cell(row=1246, column=i + 3, value=my_list6[i+167])
    for i in range(len(my_list7[167:198])):
        sheet6.cell(row=1247, column=i + 3, value=my_list7[i+167])

    #Anti acne
    for i in range(len(my_list21[167:198])):
        sheet6.cell(row=1194, column=i + 3, value=my_list21[i+167])
    for i in range(len(my_list31[167:198])):
        sheet6.cell(row=1195, column=i + 3, value=my_list31[i+167])
    for i in range(len(my_list41[167:198])):
        sheet6.cell(row=1196, column=i + 3, value=my_list41[i+167])
    for i in range(len(my_list51[167:198])):
        sheet6.cell(row=1197, column=i + 3, value=my_list51[i+167])
    for i in range(len(my_list61[167:198])):
        sheet6.cell(row=1198, column=i + 3, value=my_list61[i+167])
    for i in range(len(my_list71[167:198])):
        sheet6.cell(row=1199, column=i + 3, value=my_list71[i+167])

    #setsuko tonic
    for i in range(len(my_list211[167:198])):
        sheet6.cell(row=1226, column=i + 3, value=my_list211[i+167])
    for i in range(len(my_list311[167:198])):
        sheet6.cell(row=1227, column=i + 3, value=my_list311[i+167])
    for i in range(len(my_list411[167:198])):
        sheet6.cell(row=1228, column=i + 3, value=my_list411[i+167])
    for i in range(len(my_list511[167:198])):
        sheet6.cell(row=1229, column=i + 3, value=my_list511[i+167])
    for i in range(len(my_list611[167:198])):
        sheet6.cell(row=1230, column=i + 3, value=my_list611[i+167])
    for i in range(len(my_list711[167:198])):
        sheet6.cell(row=1231, column=i + 3, value=my_list711[i+167])

    #Anti age
    for i in range(len(my_list2111[167:198])):
        sheet6.cell(row=1210, column=i + 3, value=my_list2111[i+167])
    for i in range(len(my_list3111[167:198])):
        sheet6.cell(row=1211, column=i + 3, value=my_list3111[i+167])
    for i in range(len(my_list4111[167:198])):
        sheet6.cell(row=1212, column=i + 3, value=my_list4111[i+167])
    for i in range(len(my_list5111[167:198])):
        sheet6.cell(row=1213, column=i + 3, value=my_list5111[i+167])
    for i in range(len(my_list6111[167:198])):
        sheet6.cell(row=1214, column=i + 3, value=my_list6111[i+167])
    for i in range(len(my_list7111[167:198])):
        sheet6.cell(row=1215, column=i + 3, value=my_list7111[i+167])

    wb5.save("sheet6.xlsx")

    def get_column_label(i):
        if i < 26:
            return chr(65 + i)
        else:
            div = i // 26
            mod = i % 26
            if mod == 0:
                return get_column_label(div - 1) + 'Z'
            else:
                return get_column_label(div - 1) + get_column_label(mod)

    def CopyFromExcInGsh5():
        client = gspread.authorize(credentials)

        spreadsheet = client.open_by_key(KEY_TABLE)
        worksheet = spreadsheet.worksheet('Аналитика и статистика все компании')

        df = pd.read_excel("sheet6.xlsx")
        data_list = df.values.tolist()
        num_cols = len(data_list[0])

        cell_list = worksheet.range('A1:' + get_column_label(num_cols - 1) + str(len(data_list)))
        for cell in cell_list:
            row = (cell.row - 1) if (cell.row - 1) < len(data_list) else -1
            col = (cell.col - 1) if (cell.col - 1) < num_cols else -1
            if row != -1 and col != -1:
                value = data_list[row][col]
                if pd.notna(value):
                    cell.value = str(value)

        worksheet.update_cells(cell_list)
        print("Данные загружены")

    CopyFromExcInGsh5()